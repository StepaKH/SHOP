import openpyxl
import time


# def check_file_lock(filename):
#     try:
#         handle = win32file.CreateFile(
#             filename,
#             win32con.GENERIC_READ | win32con.GENERIC_WRITE,
#             0,
#             None,
#             win32con.OPEN_EXISTING,
#             0,
#             0
#         )
#         win32file.CloseHandle(handle)
#         return False
#     except Exception as e:
#         return True


def check_card_status(filename, phone_number):
    wb = openpyxl.load_workbook(filename, read_only=True, keep_vba=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
            if str(row[9]) == phone_number:
                return 1
        return 0
    finally:
        wb.close()


def create_card(filename, fio, phone, birthday):
    wb = openpyxl.load_workbook(filename, read_only=False)
    try:
        ws = wb.active
        last_row = ws.max_row

        fio = fio.split()

        ws.cell(row=last_row + 1, column=1).value = ws.cell(row=last_row, column=1).value
        ws.cell(row=last_row + 1, column=2).value = int(ws.cell(row=last_row, column=2).value + 1)
        ws.cell(row=last_row + 1, column=2).number_format = '0'
        ws.cell(row=last_row + 1, column=3).value = fio[0]
        ws.cell(row=last_row + 1, column=4).value = fio[1]
        ws.cell(row=last_row + 1, column=5).value = fio[2]
        ws.cell(row=last_row + 1, column=6).value = birthday
        ws.cell(row=last_row + 1, column=10).value = phone
    finally:
        wb.save(filename)
        wb.close()

# create_card('cards.xlsx', 'Хрушков Степан Игоревич', "79965677951", "10.01.2005")